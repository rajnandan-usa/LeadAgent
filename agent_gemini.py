"""
╔═══════════════════════════════════════════════════════════╗
║         RAJNANDAN AI FREELANCE AGENT                     ║
║         Powered by Google Gemini API (FREE!)             ║
╚═══════════════════════════════════════════════════════════╝
"""

import google.generativeai as genai
import json, sqlite3, smtplib, requests, openpyxl, time, os, re
from datetime import datetime, timedelta
from email.mime.multipart import MIMEMultipart
from email.mime.text import MIMEText
from openpyxl.styles import Font, PatternFill, Alignment

GEMINI_API_KEY = os.environ.get("GEMINI_API_KEY", "your_gemini_key")
GOOGLE_API_KEY = os.environ.get("GOOGLE_API_KEY", "your_places_key")
EMAIL_HOST     = os.environ.get("EMAIL_HOST", "smtp.hostinger.com")
EMAIL_PORT     = int(os.environ.get("EMAIL_PORT", "587"))
EMAIL_USER     = os.environ.get("EMAIL_USER", "hello@rajnandanweb.com")
EMAIL_PASS     = os.environ.get("EMAIL_PASS", "your_password")
DB_FILE        = "agent_leads.db"
PORTFOLIO  = "https://rajnandanweb.com"
PHONE      = "+91 7903292066"
WHATSAPP   = "917903292066"
MAX_EMAILS = 25

TARGETS = [
    ("restaurant","Kathmandu","NP"), ("travel agency","Kathmandu","NP"),
    ("hotel","Pokhara","NP"), ("restaurant","Patna","IN"),
    ("travel agency","Patna","IN"), ("clinic","Patna","IN"),
    ("school","Patna","IN"), ("gym","Muzaffarpur","IN"),
]

genai.configure(api_key=GEMINI_API_KEY)
model = genai.GenerativeModel("gemini-1.5-flash")

def ask_gemini(prompt):
    try:
        return model.generate_content(prompt).text.strip()
    except Exception as e:
        print(f"  Gemini error: {e}")
        return ""

def init_db():
    conn = sqlite3.connect(DB_FILE)
    c = conn.cursor()
    c.execute('''CREATE TABLE IF NOT EXISTS leads (
        id INTEGER PRIMARY KEY AUTOINCREMENT,
        name TEXT, category TEXT, city TEXT, country TEXT,
        phone TEXT, email TEXT DEFAULT "", website TEXT DEFAULT "",
        has_website INTEGER DEFAULT 0, rating REAL DEFAULT 0,
        place_id TEXT UNIQUE, priority TEXT DEFAULT "medium",
        status TEXT DEFAULT "new", ai_analysis TEXT DEFAULT "",
        email_sent_at TEXT, followup1_at TEXT, followup2_at TEXT,
        replied INTEGER DEFAULT 0, converted INTEGER DEFAULT 0,
        created_at TEXT DEFAULT CURRENT_TIMESTAMP)''')
    c.execute('''CREATE TABLE IF NOT EXISTS email_logs (
        id INTEGER PRIMARY KEY AUTOINCREMENT,
        lead_id INTEGER, email_type TEXT, subject TEXT,
        sent_at TEXT, status TEXT)''')
    conn.commit(); conn.close()

def dbq(q, p=(), fetch=False):
    conn = sqlite3.connect(DB_FILE)
    c = conn.cursor(); c.execute(q, p)
    r = c.fetchall() if fetch else None
    conn.commit(); conn.close()
    return r

def log(action, detail=""):
    print(f"  [{datetime.now().strftime('%H:%M:%S')}] {action}: {detail}")

def search_leads(keyword, city, country):
    try:
        r = requests.get("https://maps.googleapis.com/maps/api/place/textsearch/json",
            params={"query": f"{keyword} in {city}", "key": GOOGLE_API_KEY, "region": country.lower()},
            timeout=10).json().get("results", [])[:10]
        leads = []
        for p in r:
            d = requests.get("https://maps.googleapis.com/maps/api/place/details/json",
                params={"place_id": p["place_id"],
                        "fields": "name,formatted_phone_number,website,rating",
                        "key": GOOGLE_API_KEY}, timeout=10).json().get("result", {})
            time.sleep(0.2)
            leads.append({"name": d.get("name", p.get("name","")),
                "phone": d.get("formatted_phone_number",""), "website": d.get("website",""),
                "rating": p.get("rating", 0), "place_id": p["place_id"],
                "has_website": bool(d.get("website")),
                "city": city, "country": country, "category": keyword})
        return leads
    except Exception as e:
        print(f"  Search error: {e}"); return []

def get_email_from_site(url):
    if not url: return ""
    try:
        r = requests.get(url, headers={"User-Agent":"Mozilla/5.0"}, timeout=5)
        emails = re.findall(r'[a-zA-Z0-9._%+\-]+@[a-zA-Z0-9.\-]+\.[a-zA-Z]{2,}', r.text)
        valid = [e for e in emails if not any(e.endswith(x) for x in ['.png','.jpg','.svg','.js','.css'])]
        return valid[0] if valid else ""
    except: return ""

def analyze(lead):
    prompt = f"""Analyze this business lead for a freelance web developer. Return JSON only (no markdown):
Business: {lead['name']}, Type: {lead['category']}, City: {lead['city']} {lead['country']}
Has Website: {lead['has_website']}, Rating: {lead.get('rating', 'N/A')}
Return: {{"priority":"high/medium/low","pain_point":"one sentence","should_contact":true/false}}
High priority = no website OR rating < 3.5"""
    try:
        r = ask_gemini(prompt)
        return json.loads(re.sub(r'```json|```', '', r).strip())
    except:
        return {"priority": "high" if not lead["has_website"] else "medium",
                "pain_point": "No online presence" if not lead["has_website"] else "Needs better website",
                "should_contact": True}

def write_email(lead, analysis, etype="initial"):
    country = lead.get("country", "IN")
    lang = "Hinglish (Hindi+English mix)" if country == "IN" else "professional English"
    prompts = {
        "initial": f"Write a cold outreach email. Mention their pain point: {analysis['pain_point']}.",
        "followup1": "Write a gentle 3-day follow-up. Ask one simple question.",
        "followup2": "Write a final 7-day follow-up. Very short. Leave door open."
    }
    prompt = f"""{prompts.get(etype, prompts["initial"])}
For: {lead['name']} ({lead['category']} in {lead['city']})
Use: {lang}. Max 120 words. End with WhatsApp: https://wa.me/{WHATSAPP}
My portfolio: {PORTFOLIO} | Built: Ongofix.com, FasalVision (Play Store), Fleeta2z
Return JSON only: {{"subject":"...","body":"..."}}"""
    try:
        r = ask_gemini(prompt)
        return json.loads(re.sub(r'```json|```', '', r).strip())
    except:
        return {"subject": f"Quick suggestion for {lead['name']}",
                "body": f"Hi,\n\nI noticed {analysis['pain_point']}. I can help!\n\nPortfolio: {PORTFOLIO}\nWA: https://wa.me/{WHATSAPP}\n\nRajnandan | {PHONE}"}

def send_email(to, subject, body, name=""):
    try:
        msg = MIMEMultipart(); msg["From"] = f"Rajnandan Kushwaha <{EMAIL_USER}>"
        msg["To"] = to; msg["Subject"] = subject
        msg.attach(MIMEText(body, "plain", "utf-8"))
        with smtplib.SMTP(EMAIL_HOST, EMAIL_PORT) as s:
            s.ehlo(); s.starttls(); s.login(EMAIL_USER, EMAIL_PASS)
            s.sendmail(EMAIL_USER, to, msg.as_string())
        log("EMAIL_SENT", f"✅ {name} → {to}"); return True
    except Exception as e:
        log("EMAIL_FAIL", f"❌ {name}: {e}"); return False

def today_emails():
    t = datetime.now().date().isoformat()
    r = dbq(f"SELECT COUNT(*) FROM email_logs WHERE sent_at LIKE '{t}%' AND status='sent'", fetch=True)
    return r[0][0] if r else 0

def export_excel():
    leads = dbq("SELECT * FROM leads ORDER BY priority DESC, created_at DESC", fetch=True)
    wb = openpyxl.Workbook(); ws = wb.active; ws.title = "All Leads"
    headers = ["ID","Name","Category","City","Country","Phone","Email","Website",
               "Has Website","Rating","Priority","Status","AI Analysis",
               "Email Sent","Followup1","Followup2","Replied","Converted","Created"]
    for i, h in enumerate(headers, 1):
        c = ws.cell(row=1, column=i, value=h)
        c.font = Font(bold=True, color="FFFFFF", name="Arial")
        c.fill = PatternFill("solid", start_color="1F3864")
        c.alignment = Alignment(horizontal="center")
    for ri, row in enumerate(leads, 2):
        for ci, val in enumerate(row, 1):
            cell = ws.cell(row=ri, column=ci, value=val)
            cell.font = Font(name="Arial", size=9)
            cell.fill = PatternFill("solid", start_color="EEF2FF" if ri%2==0 else "FFFFFF")
            if ci == 11:
                colors = {"high":"C00000","medium":"E97132","low":"538135"}
                if val in colors: cell.font = Font(name="Arial", size=9, bold=True, color=colors[val])
    ws.freeze_panes = "B2"
    fname = f"leads_{datetime.now().strftime('%Y%m%d_%H%M')}.xlsx"
    wb.save(fname); log("EXCEL", fname); return fname

def run_agent():
    init_db()
    print("\n" + "="*55)
    print("🤖 RAJNANDAN AI AGENT — Powered by Gemini")
    print(f"   {datetime.now().strftime('%Y-%m-%d %H:%M')}"); print("="*55)

    sent = today_emails()
    saved = emailed = 0

    # Phase 1: Search leads
    print("\n🔍 Searching leads...")
    for keyword, city, country in TARGETS[:4]:
        print(f"  {keyword} in {city}...")
        for lead in search_leads(keyword, city, country):
            if not lead.get("email") and lead.get("website"):
                lead["email"] = get_email_from_site(lead["website"])
            an = analyze(lead)
            if not an.get("should_contact") or an["priority"] == "low": continue
            try:
                dbq('''INSERT OR IGNORE INTO leads
                    (name,category,city,country,phone,email,website,has_website,rating,place_id,priority,ai_analysis)
                    VALUES (?,?,?,?,?,?,?,?,?,?,?,?)''',
                    (lead["name"],lead["category"],lead["city"],lead["country"],
                     lead.get("phone",""),lead.get("email",""),lead.get("website",""),
                     int(lead.get("has_website",False)),lead.get("rating",0),
                     lead["place_id"],an["priority"],json.dumps(an)))
                saved += 1
                icon = "🔴" if an["priority"]=="high" else "🟡"
                print(f"    {icon} {lead['name']} [{an['priority']}]")
            except: pass
            time.sleep(0.3)

    # Phase 2: Send initial emails
    print(f"\n📧 Sending emails (limit: {MAX_EMAILS})...")
    new_leads = dbq('''SELECT id,name,category,city,country,email,has_website
                       FROM leads WHERE status="new" AND email!="" AND email_sent_at IS NULL
                       ORDER BY priority DESC LIMIT ?''', (MAX_EMAILS-sent,), fetch=True) or []
    for row in new_leads:
        if sent >= MAX_EMAILS: break
        lid,name,cat,city,country,email,has_web = row
        lead = {"name":name,"category":cat,"city":city,"country":country,"has_website":bool(has_web)}
        an_raw = dbq("SELECT ai_analysis FROM leads WHERE id=?", (lid,), fetch=True)
        an = json.loads(an_raw[0][0]) if an_raw else {"pain_point":"No website"}
        ed = write_email(lead, an, "initial")
        if send_email(email, ed["subject"], ed["body"], name):
            now = datetime.now().isoformat()
            dbq("UPDATE leads SET status='contacted',email_sent_at=?,followup1_at=?,followup2_at=? WHERE id=?",
                (now,(datetime.now()+timedelta(days=3)).date().isoformat(),
                 (datetime.now()+timedelta(days=7)).date().isoformat(),lid))
            dbq("INSERT INTO email_logs (lead_id,email_type,subject,sent_at,status) VALUES (?,?,?,?,?)",
                (lid,"initial",ed["subject"],now,"sent"))
            sent += 1; emailed += 1; time.sleep(4)

    # Phase 3: Follow-ups
    print("\n🔄 Follow-ups...")
    today = datetime.now().date().isoformat()
    for fcol, ftype in [("followup1_at","followup1"),("followup2_at","followup2")]:
        if sent >= MAX_EMAILS: break
        due = dbq(f'''SELECT id,name,category,city,country,email,has_website FROM leads
                      WHERE {fcol}<=? AND replied=0 AND status="contacted" AND email!=""
                      AND id NOT IN (SELECT lead_id FROM email_logs WHERE email_type=?) LIMIT ?''',
                  (today,ftype,MAX_EMAILS-sent), fetch=True) or []
        print(f"  {ftype}: {len(due)} due")
        for row in due:
            lid,name,cat,city,country,email,has_web = row
            lead = {"name":name,"category":cat,"city":city,"country":country,"has_website":bool(has_web)}
            ed = write_email(lead, {"pain_point":"Following up"}, ftype)
            if send_email(email, ed["subject"], ed["body"], f"[FU] {name}"):
                dbq("INSERT INTO email_logs (lead_id,email_type,subject,sent_at,status) VALUES (?,?,?,?,?)",
                    (lid,ftype,ed["subject"],datetime.now().isoformat(),"sent"))
                sent += 1; emailed += 1; time.sleep(4)

    # Phase 4: Excel
    print("\n📊 Updating Excel..."); excel_file = export_excel()

    print("\n" + "="*55 + "\n✅ DONE!")
    print(f"  Leads saved today : {saved}")
    print(f"  Emails sent today : {emailed}")
    print(f"  Excel file        : {excel_file}")
    print("="*55)

if __name__ == "__main__":
    import sys
    cmd = sys.argv[1] if len(sys.argv)>1 else "run"
    if cmd=="run": run_agent()
    elif cmd=="stats":
        init_db()
        print(f"Total: {dbq('SELECT COUNT(*) FROM leads',fetch=True)[0][0]}")
        print(f"Converted: {dbq('SELECT COUNT(*) FROM leads WHERE converted=1',fetch=True)[0][0]}")
    elif cmd=="won" and len(sys.argv)>2:
        init_db(); dbq("UPDATE leads SET converted=1,status='converted' WHERE id=?", (sys.argv[2],))
        print(f"💰 Lead #{sys.argv[2]} WON!")
    elif cmd=="excel": init_db(); print(export_excel())
